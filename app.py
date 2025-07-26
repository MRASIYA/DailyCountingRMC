from flask import Flask, render_template, request, redirect, url_for, flash, jsonify
import pandas as pd
import os
from datetime import datetime
from openpyxl import load_workbook

app = Flask(__name__)
app.secret_key = 'your-secret-key-here'

# Configuration
EXCEL_FILE = 'ISSUES.xlsx'
TRANSACTIONS_SHEET = 'Transactions'
MAIN_SHEET = 'Sheet1'

def load_materials():
    """Load materials from Excel file"""
    try:
        # Read materials from the main sheet, skip the header rows
        df = pd.read_excel(EXCEL_FILE, sheet_name=MAIN_SHEET, skiprows=1)
        
        # Find the columns we need (Material Code in column 1, Material Name in column 2)
        materials = []
        for index, row in df.iterrows():
            if pd.notna(row.iloc[1]) and pd.notna(row.iloc[2]):  # Material Code and Material Name
                material_code = str(row.iloc[1]).strip()
                material_name = str(row.iloc[2]).strip()
                if material_code != 'Material Code' and material_name != 'Material Name' and material_code != 'nan':
                    materials.append({
                        'code': material_code,
                        'name': material_name
                    })
        
        return materials
    except Exception as e:
        print(f"Error loading materials: {e}")
        return []

def save_transaction(material_code, material_name, transaction_type, quantity):
    """Save and update quantities in Excel file"""
    import time
    import shutil
    
    max_retries = 3
    retry_delay = 1
    
    for attempt in range(max_retries):
        try:
            # Create a backup first
            backup_file = f"{EXCEL_FILE}.backup"
            if os.path.exists(EXCEL_FILE):
                shutil.copy2(EXCEL_FILE, backup_file)
            
            # Load the workbook
            wb = load_workbook(EXCEL_FILE)
            ws = wb[MAIN_SHEET]
            
            # Find the row with the matching material code
            for row in ws.iter_rows(min_row=3, max_row=ws.max_row):
                if str(row[1].value) == str(material_code):
                    # Determine the column based on transaction type (0-indexed)
                    if transaction_type == 'Issues' or transaction_type == 'Issued':
                        col_idx = 6  # Issued Column (0-indexed)
                    elif transaction_type == 'Received':
                        col_idx = 5  # Received Column (0-indexed)
                    elif transaction_type == 'Return':
                        col_idx = 7  # Return Column (0-indexed)
                    else:
                        print(f"Unknown transaction type: {transaction_type}")
                        continue
                    
                    # Update the quantity
                    current_value = row[col_idx].value
                    if current_value is None:
                        current_value = 0
                    row[col_idx].value = float(current_value) + float(quantity)
                    
                    break
            
            # Save the workbook
            wb.save(EXCEL_FILE)
            wb.close()
            
            # Remove backup if successful
            if os.path.exists(backup_file):
                os.remove(backup_file)
            
            return  # Success, exit the retry loop
            
        except PermissionError as e:
            print(f"Permission error on attempt {attempt + 1}: {e}")
            if attempt < max_retries - 1:
                time.sleep(retry_delay)
                retry_delay *= 2  # Exponential backoff
            else:
                # Try to restore backup
                if os.path.exists(backup_file):
                    shutil.copy2(backup_file, EXCEL_FILE)
                    os.remove(backup_file)
                raise Exception(f"Could not save to Excel file after {max_retries} attempts. File may be open in another application.")
        except Exception as e:
            print(f"Error saving transaction on attempt {attempt + 1}: {e}")
            if attempt < max_retries - 1:
                time.sleep(retry_delay)
                retry_delay *= 2
            else:
                # Try to restore backup
                if os.path.exists(backup_file):
                    shutil.copy2(backup_file, EXCEL_FILE)
                    os.remove(backup_file)
                raise e

def get_current_stock():
    """Get current stock levels from Sheet1"""
    try:
        # Read from Sheet1, skip header rows
        df = pd.read_excel(EXCEL_FILE, sheet_name=MAIN_SHEET, skiprows=2, header=0)
        
        # Convert to list of dictionaries
        stock_data = []
        for _, row in df.iterrows():
            if pd.notna(row.iloc[1]) and pd.notna(row.iloc[2]):  # Material Code and Name exist
                material_code = str(row.iloc[1]).strip()
                material_name = str(row.iloc[2]).strip()
                
                if material_code != 'Material Code' and material_name != 'Material Name':
                    received = row.iloc[5] if pd.notna(row.iloc[5]) else 0
                    issued = row.iloc[6] if pd.notna(row.iloc[6]) else 0
                    returned = row.iloc[7] if pd.notna(row.iloc[7]) else 0
                    
                    stock_data.append({
                        'material_code': material_code,
                        'material_name': material_name,
                        'received': float(received),
                        'issued': float(issued),
                        'returned': float(returned)
                    })
        
        return stock_data
        
    except Exception as e:
        print(f"Error loading current stock: {e}")
        return []

@app.route('/')
def index():
    """Main page with form"""
    materials = load_materials()
    stock_data = get_current_stock()
    return render_template('index.html', materials=materials, stock_data=stock_data)

@app.route('/submit_transaction', methods=['GET', 'POST'])
def submit_transaction():
    """Handle form submission"""
    # Handle GET requests with a helpful message
    if request.method == 'GET':
        flash('Please use the form to submit transactions', 'info')
        return redirect(url_for('index'))
    
    # Handle POST requests (form submissions)
    try:
        material_selection = request.form['material']
        transaction_type = request.form['transaction_type']
        quantity = float(request.form['quantity'])
        
        # Parse material selection (format: "code|name")
        material_code, material_name = material_selection.split('|', 1)
        
        # Validate inputs
        if not material_code or not material_name or not transaction_type or quantity <= 0:
            flash('Please fill all fields with valid values', 'error')
            return redirect(url_for('index'))
        
        # Save transaction
        save_transaction(material_code, material_name, transaction_type, quantity)
        
        flash(f'Updated {transaction_type}: {quantity} units of {material_name} in ISSUES.xlsx', 'success')
        
    except Exception as e:
        flash(f'Error: {str(e)}', 'error')
    
    return redirect(url_for('index'))

@app.route('/api/materials')
def api_materials():
    """API endpoint to get materials"""
    return jsonify(load_materials())

@app.route('/api/stock')
def api_stock():
    """API endpoint to get current stock levels"""
    return jsonify(get_current_stock())

@app.route('/download')
def download_excel():
    """Download the ISSUES.xlsx file"""
    from flask import send_file, abort
    try:
        abs_path = os.path.abspath(EXCEL_FILE)
        if not os.path.exists(abs_path):
            print(f"Download failed, file not found at: {abs_path}")
            abort(404, description="File not found")
        return send_file(
            abs_path, 
            as_attachment=True, 
            download_name='ISSUES.xlsx',
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
    except Exception as e:
        print(f"Error during file download: {e}")
        flash(f'Error downloading file: {str(e)}', 'error')
        return redirect(url_for('index'))

if __name__ == '__main__':
    app.run(debug=True, host='0.0.0.0', port=5000)
